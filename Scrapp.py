from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
import pandas as pd
import openpyxl
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

def agrupar_entidades(df):
    # Este DataFrame agrupado almacenará los resultados finales
    df_agrupado = pd.DataFrame(columns=['Entidad', 'Materia', 'Enlace'])

    # Agrupar el DataFrame por 'Entidad' y concatenar las 'Materias' y 'Enlaces'
    for entidad, group in df.groupby('Entidad'):
        # Unir las materias y enlaces con un salto de línea HTML
        materias_html = '<br>'.join(group['Materia'])
        enlaces_html = '<br>'.join([f'<a href="{link}">Ver Enlace</a>' for link in group['Enlace'].tolist()])
        
        # Añadir al DataFrame agrupado
        df_agrupado = df_agrupado.append({
            'Entidad': entidad,
            'Materia': materias_html,
            'Enlace': enlaces_html
        }, ignore_index=True)
    
    return df_agrupado

def enviar_correo(df_agrupado, destinatario, asunto):
    # Convertir el DataFrame a HTML
    html_df = df_agrupado.to_html(escape=False, index=False)

    # HTML personalizado para el cuerpo del correo
    html_correo = f"""
    <html>
        <head>
            <style>
                body {{
                    font-family: 'Arial', sans-serif;
                    margin: 10px;
                }}
                table {{
                    border-collapse: collapse;
                    width: 100%;
                }}
                th, td {{
                    border: 1px solid #dddddd;
                    text-align: left;
                    padding: 8px;
                }}
                th {{
                    background-color: #aa0404;
                    color: white;
                }}
                .footer {{
                    margin-top: 20px;
                    font-size: 0.8em;
                }}
            </style>
        </head>
        <body>
            <h2>Hechos Escenciales</h2>
            <p>Se adjuntan los hechos escenciales más importantes del día de ayer</p>
            <!-- Incluir el DataFrame HTML aquí -->
            {html_df}
            <p class="footer">Este es un correo automatizado, por favor no responda directamente.</p>
            <!-- Imagen adjunta -->
            <img src="https://gkpb.com.br/wp-content/uploads/2018/03/novo-logo-santander-fundo-vermelho.jpg" alt="Imagende ejemplo" width="200">
        </body>
    </html>
    """



    # Crea el mensaje
    mensaje = MIMEMultipart()
    mensaje['From'] = 'pablo.castro_d@outlook.com'
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto

    # Adjunta el DataFrame en HTML al correo
    mensaje.attach(MIMEText(html_correo, 'html'))

    # Configuración del servidor SMTP de Outlook
    servidor = smtplib.SMTP('smtp.office365.com', 587)
    servidor.starttls()

    # Login al servidor
    servidor.login('pablo.castro_d@outlook.com', 'telefono2708AB_')

    # Enviar el correo
    servidor.sendmail(mensaje['From'], mensaje['To'], mensaje.as_string())
    servidor.quit()

    print("Correo enviado exitosamente!")



fecha_de_hoy = time.strftime("%d/%m/%Y")
fecha_de_ayer = time.strftime("%d/%m/%Y", time.gmtime(time.time() - 86400))
ultimo_viernes = time.strftime("%d/%m/%Y", time.gmtime(time.time() - 86400*3))


def actualizar_y_agregar_a_df(archivo='hechos_esenciales.xlsx'):
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active
    
    # Crear un DataFrame vacío con las columnas específicas que vamos a utilizar
    df = pd.DataFrame(columns=['Fecha', 'Hora', 'ID', 'Entidad', 'Materia', 'Enlace'])
    
    for indice, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if fila[-1] == 'N':  # Si el valor en la columna "ENVIADO(Y/N)" es "N"
            # Cambiar el valor de "N" a "Y" en la columna "ENVIADO(Y/N)"
            hoja.cell(row=indice, column=7, value='Y')
            
            # Agregar la fila completa al DataFrame
            df = df.append(pd.Series(fila[:-1], index=df.columns), ignore_index=True)
    
    libro.save(archivo)
    libro.close()
    
    # Mantener solo las columnas "Entidad", "Materia", "Enlace" en el DataFrame
    df = df[['Entidad', 'Materia', 'Enlace']]
    
    return df


def crear_excel():
    if not os.path.exists('hechos_esenciales.xlsx'):
        archivo = 'hechos_esenciales.xlsx'
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = 'Hechos Esenciales'
        hoja.append(['Fecha', 'Hora', 'ID', 'Entidad', 'Materia', 'Enlace', 'ENVIADO(Y/N)'])
        libro.save(archivo)
        libro.close()
    else:
        print('El archivo "hechos_esenciales.xlsx" ya existe.')
    

def añadir_a_excel(datos):
    filas_agregadas = 0
    archivo = 'hechos_esenciales.xlsx'
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active
    for fila in datos:
        if fila[0] == fecha_de_ayer or fila[0] == ultimo_viernes: #Si la fecha es la de ayer o la del último viernes
            if fila[2] not in [celda.value for celda in hoja['C']]: #Si el ID no está en la columna C
                if fila[3].lower().find('banco') == -1: #Si la entidad no es un banco
                    if (fila[3].lower().find('tanner') != -1 or fila[3].lower().find('factoring') != -1) and fila[4] == ('Colocación de valores en mercados internacionales y/o nacionales'): #Si la entidad es Tanner o Factoring y la materia es colocación de valores
                        print('Agregando fila:', fila)
                        fila.append('N')
                        hoja.append(fila)
                        filas_agregadas += 1
                    else:
                        print('La entidad', fila[3], 'no cumple con los requisitos.')
                        print('-----------------------------')
                else: #Si la entidad es un banco
                        fila.append('N')
                        hoja.append(fila)
                        filas_agregadas += 1

            else:
                print('El ID', fila[2], 'ya está en el archivo.')
                print('Nombre:', fila[3])   
                print('-----------------------------')
        else:
            print('La fecha no es la de ayer o la de hoy.')
            print('ID:', fila[2])
            print('Fecha:', fila[0])
            print('-----------------------------')

    print ('Filas agregadas: ', filas_agregadas)
    libro.save(archivo)
    libro.close()
    return 
 


def accederyobtenerdf():
    print('ACCEDIENDO A CMF.....')
    service = Service()

    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-extensions')

    driver = webdriver.Chrome(service=service, options=options)
    driver.get('https://www.cmfchile.cl/portal/principal/613/w3-channel.html')

    css_selector = 'button.btn.btn-outline-secondary.btn-sm'

    WebDriverWait(driver,20)
    time.sleep(5)

    #----VENTANA INICIAL--------
    #esperar a que el elemento sea clickeable, luego hacer click
   # WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, css_selector))).click()
   # time.sleep(2)

    #----SCROL DOWN----------
    div_row_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.ntg-box-mb.animar.tab-pills-cmf")))
    driver.execute_script("arguments[0].scrollIntoView(true);",div_row_element)
    time.sleep(2)

    WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div[3]/div/div[1]/div[1]/div/div/div[1]/div/table/tbody')))
    time.sleep(2)


    #-----EXTRAER DATOS--------
    print('....EXTRAYENDO DATOS.....')
    tabla = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div[3]/div/div[1]/div[1]/div/div/div[1]/div/table')
    filas = tabla.find_elements(By.TAG_NAME, "tr")
    datos = []
    for i in range(3, len(filas)):
        elemento = filas[i].find_elements(By.TAG_NAME, "td")
        fecha_y_hora = elemento[0].text
        fecha = fecha_y_hora.split(' ')[0]
        hora = fecha_y_hora.split(' ')[1]
        id = elemento[1].text
        entidad = elemento[2].text
        materia = elemento[3].text
        enlace = elemento[1].find_element(By.TAG_NAME, "a")
        #print('fecha:',fecha)
        #print('hora:',hora)
        #print('id:',id)
        #print('entidad:',entidad)
        #print('materia:',materia)
        #print('enlace:', enlace.get_attribute('href'))
        #print('-----------------------------')
        fila = [fecha, hora, id, entidad, materia, enlace.get_attribute('href')]
        datos.append(fila)

    añadir_a_excel(datos)
    print('....FIN EXTRACCIÓN.....')
    return 0



def main():
    crear_excel()
    accederyobtenerdf()
    df = actualizar_y_agregar_a_df()
    df_agrupado = agrupar_entidades(df)
    print(df)
    print('....EXCEL ACTUALIZADO.....')
    enviar_correo(df_agrupado,'pablo.castro_d@outlook.com', 'Boletín de Hechos Escenciales')

        

if __name__ == "__main__":
    main()

